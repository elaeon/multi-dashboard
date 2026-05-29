"""UNAM Población Escolar 2025-2026 — Licenciatura y Posgrado."""
from pathlib import Path

import dash_bootstrap_components as dbc
import plotly.express as px
import plotly.graph_objects as go
import polars as pl
from dash import Dash, Input, Output, dcc, html, no_update

DATA_DIR = Path("data/unam/poblacion_escolar")

CHART_LAYOUT = dict(
    paper_bgcolor="rgba(0,0,0,0)", plot_bgcolor="rgba(0,0,0,0)",
    font_color="#CBD5E1",
    xaxis=dict(gridcolor="#334155"),
    yaxis=dict(gridcolor="#334155"),
)
CARD_STYLE = {
    "background": "#1E293B", "border": "1px solid #334155",
    "borderRadius": "8px", "padding": "16px", "textAlign": "center",
}

# ---------------------------------------------------------------------------
# Parsing helpers
# ---------------------------------------------------------------------------
ENTIDAD_PREFIXES = (
    "Facultad", "Escuela", "Centro", "Instituto", "Programa", "Unidad", "Colegio",
)
SKIP_PREFIXES = ("T O T A L", "Para mayor", "FUENTE", "a Las", "UNAM.")


def _coerce_int(v) -> int:
    if v in (None, "-", ""):
        return 0
    try:
        return int(float(v))
    except (ValueError, TypeError):
        return 0


def _load_licenciatura() -> pl.DataFrame:
    """Parse licenciatura_2025.xlsx → carrera-level rows."""
    import openpyxl

    wb = openpyxl.load_workbook(
        str(DATA_DIR / "licenciatura_2025.xlsx"), data_only=True, read_only=True
    )
    ws = wb["licenciatura"]
    rows = list(ws.iter_rows(values_only=True))

    records = []
    entidad = ""
    for row in rows[5:]:  # skip 5 header rows
        name = str(row[0]).strip() if row[0] is not None else ""
        if not name:
            continue
        if any(name.startswith(p) for p in SKIP_PREFIXES):
            continue
        if name.startswith(ENTIDAD_PREFIXES):
            entidad = name
            continue
        pi_h = _coerce_int(row[1])
        pi_m = _coerce_int(row[2])
        rei_h = _coerce_int(row[4])
        rei_m = _coerce_int(row[5])
        records.append({
            "entidad": entidad,
            "carrera": name,
            "pi_h": pi_h, "pi_m": pi_m,
            "rei_h": rei_h, "rei_m": rei_m,
        })

    return pl.DataFrame(records).with_columns([
        (pl.col("pi_h") + pl.col("rei_h")).alias("total_h"),
        (pl.col("pi_m") + pl.col("rei_m")).alias("total_m"),
        (pl.col("pi_h") + pl.col("rei_h") + pl.col("pi_m") + pl.col("rei_m")).alias("total"),
    ])


def _load_posgrado() -> pl.DataFrame:
    """Parse posgrado_2025.xlsx → programa-level rows."""
    import openpyxl

    wb = openpyxl.load_workbook(
        str(DATA_DIR / "posgrado_2025.xlsx"), data_only=True, read_only=True
    )
    ws = wb[wb.sheetnames[0]]
    rows = list(ws.iter_rows(values_only=True))

    # columns: name, pi_h, pi_m, pi_nb, pi_total, rei_h, rei_m, rei_total, grand_total
    AREA_PATTERN = (
        "CIENCIAS FÍSICO", "CIENCIAS BIOLÓGICAS", "CIENCIAS SOCIALES",
        "HUMANIDADES", "CIENCIAS DE LA SALUD", "ARTES", "INTERDISCIPLINARIA",
        "MULTIDISCIPLINARIA", "CIENCIAS DE LA",
    )
    records = []
    area = ""
    for row in rows[5:]:
        name = str(row[0]).strip() if row[0] is not None else ""
        if not name:
            continue
        if any(name.startswith(p) for p in SKIP_PREFIXES):
            continue
        if name.upper() == name and len(name) > 5:  # área header (ALL CAPS)
            area = name.title()
            continue
        # Skip Posgrado aggregate rows (they aggregate maestría + doctorado below)
        if name.startswith("Posgrado en") or name.startswith("Posgrado "):
            area_from_prog = name.replace("Posgrado en ", "").replace("Posgrado ", "")
            continue
        pi_h = _coerce_int(row[1])
        pi_m = _coerce_int(row[2])
        rei_h = _coerce_int(row[5])
        rei_m = _coerce_int(row[6])
        if pi_h + pi_m + rei_h + rei_m == 0:
            continue
        records.append({
            "area": area,
            "programa": name,
            "pi_h": pi_h, "pi_m": pi_m,
            "rei_h": rei_h, "rei_m": rei_m,
        })

    return pl.DataFrame(records).with_columns([
        (pl.col("pi_h") + pl.col("rei_h")).alias("total_h"),
        (pl.col("pi_m") + pl.col("rei_m")).alias("total_m"),
        (pl.col("pi_h") + pl.col("rei_h") + pl.col("pi_m") + pl.col("rei_m")).alias("total"),
    ])


# ---------------------------------------------------------------------------
# Module-level data load
# ---------------------------------------------------------------------------
df_lic = _load_licenciatura()
df_pos = _load_posgrado()

# Entidad aggregates
df_lic_ent = (
    df_lic
    .group_by("entidad")
    .agg([
        pl.col("pi_h").sum(), pl.col("pi_m").sum(),
        pl.col("rei_h").sum(), pl.col("rei_m").sum(),
        pl.col("total_h").sum(), pl.col("total_m").sum(),
        pl.col("total").sum(),
    ])
    .with_columns(
        (pl.col("total_m") / pl.col("total") * 100).alias("pct_m")
    )
    .sort("total", descending=True)
)

# Area aggregates for posgrado
df_pos_area = (
    df_pos
    .group_by("area")
    .agg([
        pl.col("total_h").sum(), pl.col("total_m").sum(), pl.col("total").sum(),
    ])
    .with_columns(
        (pl.col("total_m") / pl.col("total") * 100).alias("pct_m")
    )
    .sort("total", descending=True)
)

# Historical primer ingreso time series (2007-2025)
_df_pi_raw = pl.read_csv("data/unam_egresos_y_titulacion/poblacion_escolar_lic.csv")

df_pi_trend = (
    _df_pi_raw
    .group_by("año")
    .agg(pl.col("pi_h").sum(), pl.col("pi_m").sum())
    .with_columns(
        (pl.col("pi_h") + pl.col("pi_m")).alias("pi_total"),
        (pl.col("pi_m") / (pl.col("pi_h") + pl.col("pi_m")) * 100).alias("pct_m"),
    )
    .sort("año")
)

# Carrera × year pivot for YoY trending chart
_by_year_carrera = (
    _df_pi_raw
    .group_by(["año", "carrera"])
    .agg((pl.col("pi_h") + pl.col("pi_m")).sum().alias("pi_total"))
)
_pi_pivot = _by_year_carrera.pivot(on="año", index="carrera", values="pi_total")
TREND_YEARS = sorted(
    int(c) for c in _pi_pivot.columns if c != "carrera"
)  # 2007-2025

# Posgrado programa gender gap — min 20 students
df_pos_gap = (
    df_pos
    .filter(pl.col("total") >= 20)
    .with_columns(
        (pl.col("total_m") / pl.col("total") * 100).alias("pct_m"),
    )
    .sort("pct_m")
)

# Carrera-level PI vs REI gender retention (min 30 PI / 50 REI students)
df_carrera_retention = (
    df_lic
    .group_by("carrera")
    .agg(pl.col("pi_h").sum(), pl.col("pi_m").sum(),
         pl.col("rei_h").sum(), pl.col("rei_m").sum())
    .with_columns(
        (pl.col("pi_h") + pl.col("pi_m")).alias("pi_total"),
        (pl.col("rei_h") + pl.col("rei_m")).alias("rei_total"),
    )
    .with_columns(
        (pl.col("pi_m") / pl.col("pi_total") * 100).alias("pct_m_pi"),
        (pl.col("rei_m") / pl.col("rei_total") * 100).alias("pct_m_rei"),
    )
    .with_columns(
        (pl.col("pct_m_rei") - pl.col("pct_m_pi")).alias("delta")
    )
    .filter((pl.col("pi_total") >= 30) & (pl.col("rei_total") >= 50))
    .sort("delta")
)

# Carrera gender gap — aggregate across all entidades, min 50 students total
df_gap = (
    df_lic
    .group_by("carrera")
    .agg(pl.col("total_h").sum(), pl.col("total_m").sum(), pl.col("total").sum())
    .filter(pl.col("total") >= 50)
    .with_columns(
        (pl.col("total_m") / pl.col("total") * 100).alias("pct_m"),
    )
    .sort("pct_m")
)

# ---------------------------------------------------------------------------
# Figure factories
# ---------------------------------------------------------------------------

def fig_entidad_enrollment(top_n: int) -> go.Figure:
    """Enrollment by entidad — horizontal bar sorted descending."""
    d = df_lic_ent.head(top_n)
    n = len(d)
    height = max(320, n * 28 + 80)
    fig = go.Figure()
    fig.add_trace(go.Bar(
        x=d["total_h"].to_list(),
        y=d["entidad"].to_list(),
        orientation="h",
        name="Hombres",
        marker_color="#2E86AB",
        hovertemplate="<b>%{y}</b><br>Hombres: %{x:,}<extra></extra>",
    ))
    fig.add_trace(go.Bar(
        x=d["total_m"].to_list(),
        y=d["entidad"].to_list(),
        orientation="h",
        name="Mujeres",
        marker_color="#E84855",
        hovertemplate="<b>%{y}</b><br>Mujeres: %{x:,}<extra></extra>",
    ))
    fig.update_layout(
        barmode="stack",
        title="Matrícula por entidad académica",
        height=height,
        xaxis=dict(gridcolor="#334155"),
        yaxis=dict(autorange="reversed", gridcolor="#334155", automargin=True),
        legend=dict(orientation="h", y=1.02, x=0),
        margin=dict(t=50, b=30, l=0, r=10),
        paper_bgcolor="rgba(0,0,0,0)", plot_bgcolor="rgba(0,0,0,0)",
        font_color="#CBD5E1",
    )
    return fig


def fig_gender_dist_entidad() -> go.Figure:
    """Stacked 100% bar — % mujeres per entidad, sorted by pct_m."""
    d = df_lic_ent.sort("pct_m")
    n = len(d)
    height = max(320, n * 22 + 80)

    pct_m = d["pct_m"].to_list()
    pct_h = [100 - p for p in pct_m]
    entidades = d["entidad"].to_list()
    totals = d["total"].to_list()

    fig = go.Figure()
    fig.add_trace(go.Bar(
        x=pct_h, y=entidades, orientation="h", name="Hombres",
        marker_color="#2E86AB",
        customdata=d["total_h"].to_list(),
        hovertemplate="<b>%{y}</b><br>Hombres: %{x:.1f}%  (n=%{customdata:,})<extra></extra>",
        text=[f"{p:.0f}%" if p > 8 else "" for p in pct_h],
        textposition="inside", insidetextanchor="middle",
    ))
    fig.add_trace(go.Bar(
        x=pct_m, y=entidades, orientation="h", name="Mujeres",
        marker_color="#E84855",
        customdata=d["total_m"].to_list(),
        hovertemplate="<b>%{y}</b><br>Mujeres: %{x:.1f}%  (n=%{customdata:,})<extra></extra>",
        text=[f"{p:.0f}%" if p > 8 else "" for p in pct_m],
        textposition="inside", insidetextanchor="middle",
    ))
    fig.add_shape(
        type="line", x0=50, x1=50, y0=-0.5, y1=n - 0.5,
        line=dict(color="#F4A261", width=1.5, dash="dot"),
    )
    fig.update_layout(
        barmode="stack",
        title="Distribución por género — entidades (ordenado por % mujeres)",
        height=height,
        xaxis=dict(range=[0, 100], visible=False, gridcolor="#334155"),
        yaxis=dict(gridcolor="rgba(0,0,0,0)", automargin=True),
        legend=dict(orientation="h", y=1.02, x=0),
        margin=dict(t=50, b=20, l=0, r=10),
        paper_bgcolor="rgba(0,0,0,0)", plot_bgcolor="rgba(0,0,0,0)",
        font_color="#CBD5E1",
    )
    return fig


def fig_carrera_gap(top_n: int) -> go.Figure:
    """Diverging bar showing % mujeres per carrera — top-N extremes."""
    n_each = top_n // 2
    male_top = df_gap.head(n_each)  # most male-dominated
    female_top = df_gap.tail(n_each)  # most female-dominated

    combined = pl.concat([male_top, female_top])
    pct = combined["pct_m"].to_list()
    labels = combined["carrera"].to_list()
    totals = combined["total"].to_list()

    colors = ["#2E86AB" if p < 50 else "#E84855" for p in pct]
    dev = [p - 50 for p in pct]

    fig = go.Figure()
    fig.add_vline(x=0, line_color="#F4A261", line_dash="dot", line_width=1.5)
    fig.add_trace(go.Bar(
        x=dev,
        y=labels,
        orientation="h",
        marker_color=colors,
        customdata=[[p, t] for p, t in zip(pct, totals)],
        hovertemplate=(
            "<b>%{y}</b><br>"
            "%{customdata[0]:.1f}% mujeres<br>"
            "Total: %{customdata[1]:,}<extra></extra>"
        ),
        text=[f"{p:.0f}%" for p in pct],
        textposition="auto",
        insidetextanchor="middle",
    ))
    height = max(400, len(combined) * 26 + 100)
    # Range padded beyond max deviation so "auto" outside text has room
    max_dev = max(abs(d) for d in dev) if dev else 50
    axis_range = [-(max_dev + 12), max_dev + 12]
    fig.update_layout(
        title=f"Brecha de género por carrera — top {n_each} más extremas (mín. 50 alumnos)",
        height=height,
        xaxis=dict(
            title="Desviación del 50%",
            tickvals=[-40, -30, -20, -10, 0, 10, 20, 30, 40],
            ticktext=["10%", "20%", "30%", "40%", "50%", "60%", "70%", "80%", "90%"],
            gridcolor="#334155",
            range=axis_range,
            automargin=True,
        ),
        yaxis=dict(
            gridcolor="rgba(0,0,0,0)",
            autorange="reversed",
            automargin=True,
        ),
        margin=dict(t=50, b=40, l=0, r=0),
        paper_bgcolor="rgba(0,0,0,0)", plot_bgcolor="rgba(0,0,0,0)",
        font_color="#CBD5E1",
        showlegend=False,
    )
    return fig


def fig_posgrado_gap(top_n: int) -> go.Figure:
    """Diverging bar showing % mujeres per posgrado programa — top-N extremes."""
    n_each = top_n // 2
    combined = pl.concat([df_pos_gap.head(n_each), df_pos_gap.tail(n_each)])
    pct = combined["pct_m"].to_list()
    # Truncate long programa names
    labels = [p[:60] for p in combined["programa"].to_list()]
    totals = combined["total"].to_list()

    colors = ["#2E86AB" if p < 50 else "#E84855" for p in pct]
    dev = [p - 50 for p in pct]
    max_dev = max(abs(d) for d in dev) if dev else 50

    fig = go.Figure()
    fig.add_vline(x=0, line_color="#F4A261", line_dash="dot", line_width=1.5)
    fig.add_trace(go.Bar(
        x=dev,
        y=labels,
        orientation="h",
        marker_color=colors,
        customdata=[[p, t] for p, t in zip(pct, totals)],
        hovertemplate=(
            "<b>%{y}</b><br>"
            "%{customdata[0]:.1f}% mujeres<br>"
            "Total: %{customdata[1]:,}<extra></extra>"
        ),
        text=[f"{p:.0f}%" for p in pct],
        textposition="auto",
        insidetextanchor="middle",
    ))
    height = max(400, len(combined) * 26 + 100)
    fig.update_layout(
        title=f"Brecha de género — posgrado, top {n_each} más extremos (mín. 20 alumnos)",
        height=height,
        xaxis=dict(
            title="Desviación del 50%",
            tickvals=[-40, -30, -20, -10, 0, 10, 20, 30, 40],
            ticktext=["10%", "20%", "30%", "40%", "50%", "60%", "70%", "80%", "90%"],
            gridcolor="#334155",
            range=[-(max_dev + 12), max_dev + 12],
            automargin=True,
        ),
        yaxis=dict(gridcolor="rgba(0,0,0,0)", autorange="reversed", automargin=True),
        margin=dict(t=50, b=40, l=0, r=0),
        paper_bgcolor="rgba(0,0,0,0)", plot_bgcolor="rgba(0,0,0,0)",
        font_color="#CBD5E1",
        showlegend=False,
    )
    return fig


def fig_posgrado_area() -> go.Figure:
    """Stacked 100% bar — posgrado enrollment by área with gender split."""
    d = df_pos_area.sort("pct_m")
    pct_m = d["pct_m"].to_list()
    pct_h = [100 - p for p in pct_m]
    areas = d["area"].to_list()
    n = len(d)
    height = max(280, n * 32 + 80)

    fig = go.Figure()
    fig.add_trace(go.Bar(
        x=pct_h, y=areas, orientation="h", name="Hombres",
        marker_color="#2E86AB",
        customdata=d["total_h"].to_list(),
        hovertemplate="<b>%{y}</b><br>Hombres: %{x:.1f}%  (n=%{customdata:,})<extra></extra>",
        text=[f"{p:.0f}%" if p > 8 else "" for p in pct_h],
        textposition="inside", insidetextanchor="middle",
    ))
    fig.add_trace(go.Bar(
        x=pct_m, y=areas, orientation="h", name="Mujeres",
        marker_color="#E84855",
        customdata=d["total_m"].to_list(),
        hovertemplate="<b>%{y}</b><br>Mujeres: %{x:.1f}%  (n=%{customdata:,})<extra></extra>",
        text=[f"{p:.0f}%" if p > 8 else "" for p in pct_m],
        textposition="inside", insidetextanchor="middle",
    ))
    fig.add_shape(
        type="line", x0=50, x1=50, y0=-0.5, y1=n - 0.5,
        line=dict(color="#F4A261", width=1.5, dash="dot"),
    )
    fig.update_layout(
        barmode="stack",
        title="Posgrado: distribución por género por área de conocimiento",
        height=height,
        xaxis=dict(range=[0, 100], visible=False, gridcolor="#334155"),
        yaxis=dict(gridcolor="rgba(0,0,0,0)", automargin=True),
        legend=dict(orientation="h", y=1.02, x=0),
        margin=dict(t=50, b=20, l=0, r=10),
        paper_bgcolor="rgba(0,0,0,0)", plot_bgcolor="rgba(0,0,0,0)",
        font_color="#CBD5E1",
    )
    return fig


def fig_ingreso_vs_reingreso() -> go.Figure:
    """Compare % mujeres in primer ingreso vs reingreso across entidades."""
    d = df_lic_ent.with_columns([
        (pl.col("pi_m") / (pl.col("pi_h") + pl.col("pi_m")) * 100).alias("pct_m_pi"),
        (pl.col("rei_m") / (pl.col("rei_h") + pl.col("rei_m")) * 100).alias("pct_m_rei"),
    ]).sort("pct_m_rei")

    labels = d["entidad"].to_list()
    pi = d["pct_m_pi"].to_list()
    rei = d["pct_m_rei"].to_list()
    n = len(labels)
    height = max(320, n * 22 + 80)

    # Connector lines
    x_lines, y_lines = [], []
    for a, b, lbl in zip(pi, rei, labels):
        x_lines += [a, b, None]
        y_lines += [lbl, lbl, None]

    deltas = [b - a for a, b in zip(pi, rei)]
    colors = ["#3BB273" if d >= 0 else "#E84855" for d in deltas]

    fig = go.Figure()
    fig.add_trace(go.Scatter(
        x=x_lines, y=y_lines, mode="lines",
        line=dict(color="#475569", width=1.5), showlegend=False, hoverinfo="skip",
    ))
    fig.add_trace(go.Scatter(
        x=pi, y=labels, mode="markers", name="Primer ingreso",
        marker=dict(color="#94A3B8", size=9, symbol="circle-open",
                    line=dict(color="#94A3B8", width=2)),
        hovertemplate="<b>%{y}</b><br>Primer ingreso: %{x:.1f}%<extra></extra>",
    ))
    fig.add_trace(go.Scatter(
        x=rei, y=labels, mode="markers", name="Reingreso",
        marker=dict(color=colors, size=9),
        customdata=deltas,
        hovertemplate="<b>%{y}</b><br>Reingreso: %{x:.1f}%  (Δ %{customdata:+.1f} pp)<extra></extra>",
    ))
    fig.add_vline(x=50, line_color="#F4A261", line_dash="dot", line_width=1.5)
    fig.update_layout(
        title="% mujeres: primer ingreso vs reingreso por entidad",
        height=height,
        xaxis=dict(title="% mujeres", range=[0, 100], gridcolor="#334155"),
        yaxis=dict(gridcolor="rgba(0,0,0,0)", automargin=True),
        legend=dict(orientation="h", y=1.02, x=0),
        margin=dict(t=50, b=40, l=0, r=10),
        paper_bgcolor="rgba(0,0,0,0)", plot_bgcolor="rgba(0,0,0,0)",
        font_color="#CBD5E1",
    )
    return fig


def fig_carrera_retention(top_n: int) -> go.Figure:
    """Top-N carreras where women gain or lose share from primer ingreso to reingreso."""
    n_each = top_n // 2
    combined = pl.concat([
        df_carrera_retention.head(n_each),  # biggest drops
        df_carrera_retention.tail(n_each),  # biggest gains
    ])
    delta = combined["delta"].to_list()
    labels = combined["carrera"].to_list()
    pct_pi = combined["pct_m_pi"].to_list()
    pct_rei = combined["pct_m_rei"].to_list()
    pi_total = combined["pi_total"].to_list()
    rei_total = combined["rei_total"].to_list()

    colors = ["#3BB273" if d > 0 else "#E84855" for d in delta]
    height = max(400, len(combined) * 26 + 100)
    max_abs = max(abs(d) for d in delta) if delta else 1

    fig = go.Figure()
    fig.add_vline(x=0, line_color="#475569", line_dash="dot", line_width=1.5)
    fig.add_trace(go.Bar(
        x=delta,
        y=labels,
        orientation="h",
        marker_color=colors,
        customdata=[[d, pi, rei, pt, rt]
                    for d, pi, rei, pt, rt in zip(delta, pct_pi, pct_rei, pi_total, rei_total)],
        hovertemplate=(
            "<b>%{y}</b><br>"
            "Δ: %{customdata[0]:+.1f} pp<br>"
            "Primer ingreso: %{customdata[1]:.1f}%M  (n=%{customdata[3]:,})<br>"
            "Reingreso: %{customdata[2]:.1f}%M  (n=%{customdata[4]:,})<extra></extra>"
        ),
        text=[f"{d:+.1f} pp" for d in delta],
        textposition="auto",
        insidetextanchor="middle",
    ))
    fig.update_layout(
        title=f"Carreras donde las mujeres ganan o pierden presencia — primer ingreso → reingreso",
        height=height,
        xaxis=dict(
            title="Δ puntos porcentuales (% mujeres reingreso − % mujeres primer ingreso)",
            gridcolor="#334155",
            range=[-(max_abs * 1.3), max_abs * 1.3],
            ticksuffix=" pp",
            automargin=True,
        ),
        yaxis=dict(gridcolor="rgba(0,0,0,0)", autorange="reversed", automargin=True),
        margin=dict(t=50, b=50, l=0, r=0),
        paper_bgcolor="rgba(0,0,0,0)", plot_bgcolor="rgba(0,0,0,0)",
        font_color="#CBD5E1",
        showlegend=False,
    )
    return fig


def fig_ingreso_trend() -> go.Figure:
    """Primer ingreso total por año — líneas H, M y total con % mujeres."""
    years = df_pi_trend["año"].to_list()
    pi_h = df_pi_trend["pi_h"].to_list()
    pi_m = df_pi_trend["pi_m"].to_list()
    pi_total = df_pi_trend["pi_total"].to_list()
    pct_m = df_pi_trend["pct_m"].to_list()

    from plotly.subplots import make_subplots
    fig = make_subplots(
        rows=2, cols=1,
        row_heights=[0.65, 0.35],
        shared_xaxes=True,
        vertical_spacing=0.08,
    )

    # Panel 1: absolute numbers
    fig.add_trace(go.Scatter(
        x=years, y=pi_total, mode="lines+markers", name="Total",
        line=dict(color="#F8FAFC", width=2),
        marker=dict(size=5),
        hovertemplate="<b>%{x}</b><br>Total: %{y:,}<extra></extra>",
    ), row=1, col=1)
    fig.add_trace(go.Scatter(
        x=years, y=pi_m, mode="lines+markers", name="Mujeres",
        line=dict(color="#E84855", width=2),
        marker=dict(size=5),
        hovertemplate="<b>%{x}</b><br>Mujeres: %{y:,}<extra></extra>",
    ), row=1, col=1)
    fig.add_trace(go.Scatter(
        x=years, y=pi_h, mode="lines+markers", name="Hombres",
        line=dict(color="#2E86AB", width=2),
        marker=dict(size=5),
        hovertemplate="<b>%{x}</b><br>Hombres: %{y:,}<extra></extra>",
    ), row=1, col=1)

    # Panel 2: % mujeres
    fig.add_trace(go.Scatter(
        x=years, y=pct_m, mode="lines+markers", name="% mujeres",
        line=dict(color="#F4A261", width=2),
        marker=dict(size=5),
        hovertemplate="<b>%{x}</b><br>% mujeres: %{y:.1f}%<extra></extra>",
        showlegend=False,
    ), row=2, col=1)
    fig.add_hline(y=50, line_color="#475569", line_dash="dot", line_width=1, row=2, col=1)

    fig.update_layout(
        title="Tendencia de primer ingreso — licenciatura (2007-2025)",
        height=500,
        legend=dict(orientation="h", y=1.04, x=0),
        margin=dict(t=60, b=40, l=10, r=10),
        paper_bgcolor="rgba(0,0,0,0)", plot_bgcolor="rgba(0,0,0,0)",
        font_color="#CBD5E1",
        hovermode="x unified",
    )
    fig.update_xaxes(gridcolor="#334155", tickmode="linear", dtick=2)
    fig.update_yaxes(gridcolor="#334155")
    fig.update_yaxes(title_text="Alumnos", row=1, col=1)
    fig.update_yaxes(title_text="% mujeres", row=2, col=1)
    return fig


def fig_carrera_yoy(year: int, top_n: int = 10) -> go.Figure:
    """Top-N gaining and shrinking carreras by YoY % change in primer ingreso.

    Uses relative change (pct_change) so large and small carreras are comparable.
    Both years must have >= 50 students to filter out noise from tiny programs.
    """
    prev = str(year - 1)
    curr = str(year)
    if prev not in _pi_pivot.columns or curr not in _pi_pivot.columns:
        return go.Figure().update_layout(title=f"Sin datos para {year}", **CHART_LAYOUT)

    delta = (
        _pi_pivot
        .filter(pl.col(prev).is_not_null() & pl.col(curr).is_not_null())
        .select(["carrera", prev, curr])
        .with_columns(pl.col(prev).cast(pl.Float64), pl.col(curr).cast(pl.Float64))
        .filter((pl.col(prev) >= 50) & (pl.col(curr) >= 50))
        .with_columns(
            ((pl.col(curr) - pl.col(prev)) / pl.col(prev) * 100).alias("pct_change"),
            (pl.col(curr) - pl.col(prev)).alias("delta_abs"),
        )
        .sort("pct_change")
    )

    combined = pl.concat([delta.head(top_n), delta.tail(top_n)])
    pct = combined["pct_change"].to_list()
    labels = combined["carrera"].to_list()
    pi_curr = combined[curr].cast(pl.Int64).to_list()
    pi_prev = combined[prev].cast(pl.Int64).to_list()
    delta_abs = combined["delta_abs"].cast(pl.Int64).to_list()

    colors = ["#3BB273" if p > 0 else "#E84855" for p in pct]
    height = max(360, len(combined) * 26 + 100)
    max_abs = max(abs(p) for p in pct) if pct else 1

    fig = go.Figure()
    fig.add_vline(x=0, line_color="#475569", line_dash="dot", line_width=1.5)
    fig.add_trace(go.Bar(
        x=pct,
        y=labels,
        orientation="h",
        marker_color=colors,
        customdata=[[p, c, pr, da] for p, c, pr, da in zip(pct, pi_curr, pi_prev, delta_abs)],
        hovertemplate=(
            "<b>%{y}</b><br>"
            "Cambio: %{customdata[0]:+.1f}%  (%{customdata[3]:+,} alumnos)<br>"
            f"{year}: %{{customdata[1]:,}}<br>"
            f"{year - 1}: %{{customdata[2]:,}}<extra></extra>"
        ),
        text=[f"{p:+.1f}%" for p in pct],
        textposition="auto",
        insidetextanchor="middle",
    ))
    fig.update_layout(
        title=f"Carreras con mayor cambio relativo en primer ingreso — {year - 1} → {year}",
        height=height,
        xaxis=dict(
            title="Cambio relativo (%)",
            gridcolor="#334155",
            range=[-(max_abs * 1.3), max_abs * 1.3],
            ticksuffix="%",
            automargin=True,
        ),
        yaxis=dict(gridcolor="rgba(0,0,0,0)", autorange="reversed", automargin=True),
        margin=dict(t=50, b=40, l=0, r=0),
        paper_bgcolor="rgba(0,0,0,0)", plot_bgcolor="rgba(0,0,0,0)",
        font_color="#CBD5E1",
        showlegend=False,
    )
    return fig


# ---------------------------------------------------------------------------
# KPIs
# ---------------------------------------------------------------------------
def _kpi(value: str, label: str, sub: str = "") -> html.Div:
    return html.Div([
        html.Div(value, style={"fontSize": "2rem", "fontWeight": "700", "color": "#F8FAFC"}),
        html.Div(label, style={"fontSize": "0.85rem", "color": "#94A3B8", "marginTop": "4px"}),
        html.Div(sub, style={"fontSize": "0.75rem", "color": "#64748B"}) if sub else html.Span(),
    ], style=CARD_STYLE)


total_lic = int(df_lic["total"].sum())
pct_m_lic = float(df_lic["total_m"].sum() / total_lic * 100)
total_pos = int(df_pos["total"].sum())
pct_m_pos = float(df_pos["total_m"].sum() / total_pos * 100)
n_carreras = df_lic["carrera"].n_unique()
n_entidades = df_lic["entidad"].n_unique()

# ---------------------------------------------------------------------------
# App layout
# ---------------------------------------------------------------------------
app = Dash(
    __name__,
    external_stylesheets=[dbc.themes.DARKLY],
    title="UNAM Población Escolar",
    suppress_callback_exceptions=True,
)

app.layout = dbc.Container(
    fluid=True,
    style={"backgroundColor": "#0F172A", "minHeight": "100vh", "padding": "24px"},
    children=[
        html.H2(
            "UNAM — Población Escolar 2025-2026",
            style={"color": "#F8FAFC", "fontWeight": "700", "marginBottom": "24px"},
        ),

        # KPIs
        dbc.Row([
            dbc.Col(_kpi(f"{total_lic:,}", "Alumnos licenciatura", "escolarizada"), md=3),
            dbc.Col(_kpi(f"{pct_m_lic:.1f}%", "Mujeres en licenciatura"), md=3),
            dbc.Col(_kpi(f"{total_pos:,}", "Alumnos posgrado", "maestría y doctorado"), md=3),
            dbc.Col(_kpi(f"{pct_m_pos:.1f}%", "Mujeres en posgrado"), md=3),
        ], className="g-3 mb-4"),

        dbc.Row([
            dbc.Col(_kpi(str(n_carreras), "Carreras de licenciatura"), md=3),
            dbc.Col(_kpi(str(n_entidades), "Entidades académicas"), md=3),
        ], className="g-3 mb-4"),

        # Tabs
        dbc.Tabs(
            id="tabs",
            active_tab="tab-gap",
            style={"backgroundColor": "#0F172A"},
            children=[
                dbc.Tab(label="Brecha de género", tab_id="tab-gap",
                        label_style={"color": "#94A3B8"},
                        active_label_style={"color": "#F8FAFC", "fontWeight": "600",
                                            "borderTop": "2px solid #2E86AB"}),
                dbc.Tab(label="Matrícula por entidad", tab_id="tab-ent",
                        label_style={"color": "#94A3B8"},
                        active_label_style={"color": "#F8FAFC", "fontWeight": "600",
                                            "borderTop": "2px solid #2E86AB"}),
                dbc.Tab(label="Ingreso vs Reingreso", tab_id="tab-ingreso",
                        label_style={"color": "#94A3B8"},
                        active_label_style={"color": "#F8FAFC", "fontWeight": "600",
                                            "borderTop": "2px solid #2E86AB"}),
                dbc.Tab(label="Posgrado", tab_id="tab-pos",
                        label_style={"color": "#94A3B8"},
                        active_label_style={"color": "#F8FAFC", "fontWeight": "600",
                                            "borderTop": "2px solid #2E86AB"}),
                dbc.Tab(label="Tendencia primer ingreso", tab_id="tab-trend",
                        label_style={"color": "#94A3B8"},
                        active_label_style={"color": "#F8FAFC", "fontWeight": "600",
                                            "borderTop": "2px solid #2E86AB"}),
            ],
        ),

        html.Div(id="tab-content", style={"marginTop": "16px"}),
    ],
)


# ---------------------------------------------------------------------------
# Callbacks
# ---------------------------------------------------------------------------
@app.callback(Output("tab-content", "children"), Input("tabs", "active_tab"))
def render_tab(tab):
    if tab == "tab-gap":
        n_pos_gap = len(df_pos_gap)
        return html.Div([
            html.H6("Licenciatura", style={"color": "#94A3B8", "marginBottom": "8px"}),
            dbc.Row([
                dbc.Col([
                    html.Label("Top N carreras más extremas:",
                               style={"color": "#94A3B8", "fontSize": "0.85rem"}),
                    dcc.Slider(
                        id="gap-slider", min=10, max=60, step=10, value=30,
                        marks={v: str(v) for v in range(10, 70, 10)},
                    ),
                ], md=6),
                dbc.Col([
                    html.Label(" ", style={"display": "block", "fontSize": "0.85rem"}),
                    dbc.Button("Descargar CSV", id="btn-gap-csv",
                               color="secondary", size="sm", style={"marginTop": "4px"}),
                    dcc.Download(id="download-gap-csv"),
                ], md=2, className="d-flex flex-column justify-content-end"),
            ], className="mb-2"),
            dbc.Row([dbc.Col(dcc.Graph(id="g-gap"))], className="mb-4"),
            html.H6("Posgrado", style={"color": "#94A3B8", "marginBottom": "8px"}),
            dbc.Row([
                dbc.Col([
                    html.Label("Top N programas más extremos:",
                               style={"color": "#94A3B8", "fontSize": "0.85rem"}),
                    dcc.Slider(
                        id="pos-gap-slider", min=10, max=min(60, n_pos_gap), step=10, value=30,
                        marks={v: str(v) for v in range(10, min(70, n_pos_gap + 1), 10)},
                    ),
                ], md=6),
                dbc.Col([
                    html.Label(" ", style={"display": "block", "fontSize": "0.85rem"}),
                    dbc.Button("Descargar CSV", id="btn-pos-gap-csv",
                               color="secondary", size="sm", style={"marginTop": "4px"}),
                    dcc.Download(id="download-pos-gap-csv"),
                ], md=2, className="d-flex flex-column justify-content-end"),
            ], className="mb-2"),
            dbc.Row([dbc.Col(dcc.Graph(id="g-pos-gap"))]),
        ])
    if tab == "tab-ent":
        return html.Div([
            dbc.Row([
                dbc.Col([
                    html.Label("Mostrar top N entidades:",
                               style={"color": "#94A3B8", "fontSize": "0.85rem"}),
                    dcc.Slider(
                        id="ent-slider", min=5, max=n_entidades, step=5, value=n_entidades,
                        marks={v: str(v) for v in range(5, n_entidades + 1, 5)},
                    ),
                ], md=6),
            ], className="mb-3"),
            dbc.Row([dbc.Col(dcc.Graph(id="g-enrollment"))]),
            dbc.Row([dbc.Col(dcc.Graph(id="g-gender-ent",
                                       figure=fig_gender_dist_entidad()))]),
        ])
    if tab == "tab-ingreso":
        return html.Div([
            dbc.Row([dbc.Col(dcc.Graph(figure=fig_ingreso_vs_reingreso()))], className="mb-4"),
            html.H6("Carreras donde las mujeres ganan o pierden presencia",
                    style={"color": "#94A3B8", "marginBottom": "8px"}),
            dbc.Row([
                dbc.Col([
                    html.Label("Top N carreras más extremas:",
                               style={"color": "#94A3B8", "fontSize": "0.85rem"}),
                    dcc.Slider(
                        id="retention-slider", min=10, max=40, step=10, value=20,
                        marks={v: str(v) for v in range(10, 50, 10)},
                    ),
                ], md=6),
            ], className="mb-2"),
            dbc.Row([dbc.Col(dcc.Graph(id="g-retention"))]),
        ])
    if tab == "tab-pos":
        return dbc.Row([dbc.Col(dcc.Graph(figure=fig_posgrado_area()))])
    if tab == "tab-trend":
        year_options = [{"label": str(y), "value": y} for y in TREND_YEARS[1:]]
        return html.Div([
            dbc.Row([dbc.Col(dcc.Graph(figure=fig_ingreso_trend()))], className="mb-4"),
            html.H6("Carreras en tendencia por año", style={"color": "#94A3B8", "marginBottom": "8px"}),
            dbc.Row([
                dbc.Col([
                    html.Label("Año:", style={"color": "#94A3B8", "fontSize": "0.85rem"}),
                    dcc.Dropdown(
                        id="trend-year",
                        options=year_options,
                        value=TREND_YEARS[-1],
                        clearable=False,
                        style={"backgroundColor": "#1E293B", "color": "#CBD5E1"},
                    ),
                ], md=3),
                dbc.Col([
                    html.Label("Top N:", style={"color": "#94A3B8", "fontSize": "0.85rem"}),
                    dcc.Slider(
                        id="trend-topn", min=5, max=20, step=5, value=10,
                        marks={v: str(v) for v in range(5, 25, 5)},
                    ),
                ], md=5),
                dbc.Col([
                    html.Label(" ", style={"display": "block", "fontSize": "0.85rem"}),
                    dbc.Button(
                        "Descargar CSV", id="btn-yoy-csv",
                        color="secondary", size="sm",
                        style={"marginTop": "4px"},
                    ),
                    dcc.Download(id="download-yoy-csv"),
                ], md=2, className="d-flex flex-column justify-content-end"),
            ], className="mb-2"),
            dbc.Row([dbc.Col(dcc.Graph(id="g-carrera-yoy"))]),
        ])
    return html.Div()


@app.callback(Output("g-gap", "figure"), Input("gap-slider", "value"))
def update_gap(top_n):
    return fig_carrera_gap(top_n or 30)


@app.callback(
    Output("g-enrollment", "figure"),
    Input("ent-slider", "value"),
)
def update_enrollment(top_n):
    return fig_entidad_enrollment(top_n or n_entidades)


@app.callback(Output("g-pos-gap", "figure"), Input("pos-gap-slider", "value"))
def update_pos_gap(top_n):
    return fig_posgrado_gap(top_n or 30)


@app.callback(
    Output("g-carrera-yoy", "figure"),
    Input("trend-year", "value"),
    Input("trend-topn", "value"),
)
def update_carrera_yoy(year, top_n):
    return fig_carrera_yoy(year or TREND_YEARS[-1], top_n or 10)


@app.callback(
    Output("download-yoy-csv", "data"),
    Input("btn-yoy-csv", "n_clicks"),
    Input("trend-year", "value"),
    Input("trend-topn", "value"),
    prevent_initial_call=True,
)
def download_yoy_csv(n_clicks, year, top_n):
    from dash import ctx
    if ctx.triggered_id != "btn-yoy-csv":
        return no_update
    year = year or TREND_YEARS[-1]
    top_n = top_n or 10
    prev = str(year - 1)
    curr = str(year)
    delta = (
        _pi_pivot
        .filter(pl.col(prev).is_not_null() & pl.col(curr).is_not_null())
        .select(["carrera", prev, curr])
        .with_columns(pl.col(prev).cast(pl.Float64), pl.col(curr).cast(pl.Float64))
        .filter((pl.col(prev) >= 50) & (pl.col(curr) >= 50))
        .with_columns(
            ((pl.col(curr) - pl.col(prev)) / pl.col(prev) * 100).alias("pct_change"),
            (pl.col(curr) - pl.col(prev)).alias("delta_abs"),
        )
        .sort("pct_change")
    )
    combined = pl.concat([delta.head(top_n), delta.tail(top_n)])
    csv_str = (
        combined
        .rename({prev: f"pi_{year-1}", curr: f"pi_{year}"})
        .with_columns(pl.col(f"pi_{year-1}").cast(pl.Int64), pl.col(f"pi_{year}").cast(pl.Int64))
        .write_csv()
    )
    return dcc.send_string(csv_str, filename=f"tendencia_carreras_{year}.csv")


@app.callback(
    Output("download-gap-csv", "data"),
    Input("btn-gap-csv", "n_clicks"),
    Input("gap-slider", "value"),
    prevent_initial_call=True,
)
def download_gap_csv(n_clicks, top_n):
    from dash import ctx
    if ctx.triggered_id != "btn-gap-csv":
        return no_update
    top_n = top_n or 30
    n_each = top_n // 2
    combined = pl.concat([df_gap.head(n_each), df_gap.tail(n_each)])
    csv_str = combined.select(["carrera", "total_h", "total_m", "total", "pct_m"]).write_csv()
    return dcc.send_string(csv_str, filename=f"brecha_genero_licenciatura_top{top_n}.csv")


@app.callback(
    Output("download-pos-gap-csv", "data"),
    Input("btn-pos-gap-csv", "n_clicks"),
    Input("pos-gap-slider", "value"),
    prevent_initial_call=True,
)
def download_pos_gap_csv(n_clicks, top_n):
    from dash import ctx
    if ctx.triggered_id != "btn-pos-gap-csv":
        return no_update
    top_n = top_n or 30
    n_each = top_n // 2
    combined = pl.concat([df_pos_gap.head(n_each), df_pos_gap.tail(n_each)])
    csv_str = combined.select(["area", "programa", "total_h", "total_m", "total", "pct_m"]).write_csv()
    return dcc.send_string(csv_str, filename=f"brecha_genero_posgrado_top{top_n}.csv")


@app.callback(Output("g-retention", "figure"), Input("retention-slider", "value"))
def update_retention(top_n):
    return fig_carrera_retention(top_n or 20)


if __name__ == "__main__":
    app.run(debug=True, port=8052)
