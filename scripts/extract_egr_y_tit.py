"""Extract the four sections of the 'egr y tit' sheet from every UNAM yearbook
in data/unam/general/ and write one CSV per section into egresos_y_titulacion/.

Each yearbook file is named YYYY.xls(x). The sheet contains, in order:
  1. EGRESO
  2. EXÁMENES DE GRADO  (older years: ...Y DIPLOMAS DE ESPECIALIZACIÓN)
  3. EXÁMENES PROFESIONALES Y OTRAS OPCIONES DE TITULACIÓN
  4. TÍTULOS EXPEDIDOS

Sections are delimited by a row whose first cell is 'UNAM'.

In 2007.xls the sheet is laid out one column to the right (sections 1 and 4
keep `UNAM` and the title in column B; every data row holds categoria in
column B and valor in column C). `pick_pair` collapses that shift by taking
the first non-empty cell as categoria and the cell to its right as valor.
"""

import csv
import re
from pathlib import Path

import openpyxl
import xlrd

DATA_DIR = Path("data/unam/general")
OUT_DIR = Path("data/unam_egresos_y_titulacion")

SECTION_FILES = [
    "egreso.csv",
    "examenes_de_grado.csv",
    "examenes_profesionales.csv",
    "titulos_expedidos.csv",
]

FOOTER_PREFIX = "Para mayor información"


def read_sheet_rows(path: Path) -> list[list]:
    """Return rows of the 'egr y tit' sheet as up-to-4-cell lists."""
    if path.suffix == ".xls":
        wb = xlrd.open_workbook(str(path))
        name = next(n for n in wb.sheet_names() if "egr" in n.lower())
        ws = wb.sheet_by_name(name)
        return [[ws.cell_value(r, c) if c < ws.ncols else None for c in range(4)]
                for r in range(ws.nrows)]

    wb = openpyxl.load_workbook(str(path), data_only=True, read_only=True)
    name = next(n for n in wb.sheetnames if "egr" in n.lower())
    ws = wb[name]
    return [[row[c] if c < len(row) else None for c in range(4)]
            for row in ws.iter_rows(values_only=True)]


def pick_pair(row: list) -> tuple:
    """Return (categoria, valor) by skipping leading empty cells.

    Standard layout has categoria in col 0; 2007.xls shifts it to col 1.
    """
    for i, v in enumerate(row):
        if v is None:
            continue
        if isinstance(v, str) and not v.strip():
            continue
        nxt = row[i + 1] if i + 1 < len(row) else None
        return v, nxt
    return None, None


def split_sections(rows: list[list]) -> list[list[list]]:
    """Split rows into sections, each starting with a 'UNAM' row."""
    sections, current = [], None
    for a, b in rows:
        first = str(a).strip() if a is not None else ""
        if first.upper() == "UNAM":
            if current is not None:
                sections.append(current)
            current = []
            continue
        if current is None:
            continue
        current.append([a, b])
    if current is not None:
        sections.append(current)
    return sections


def clean_section(section_rows: list[list]) -> tuple[str, list[tuple[str, float]]]:
    """Return (section_title, list of (categoria, valor)) for one section.

    Drops the title row, an optional standalone year row, footer rows,
    and any fully blank rows.
    """
    title = ""
    data: list[tuple[str, float]] = []
    title_seen = False
    for a, b in section_rows:
        a_str = "" if a is None else str(a).strip()
        b_blank = b is None or (isinstance(b, str) and not b.strip())

        if not a_str and b_blank:
            continue
        if a_str.startswith(FOOTER_PREFIX):
            continue
        if not title_seen:
            title = a_str
            title_seen = True
            continue
        # standalone year row (e.g. 2024 with empty value column)
        if b_blank and re.fullmatch(r"\d{4}(?:\.0)?", a_str):
            continue
        data.append((a_str, b))
    return title, data


def parse_year(filename: str) -> int:
    m = re.match(r"(\d{4})", filename)
    if not m:
        raise ValueError(f"cannot parse year from {filename}")
    return int(m.group(1))


def coerce_value(v) -> str:
    if v is None:
        return ""
    if isinstance(v, float) and v.is_integer():
        return str(int(v))
    return str(v).strip()


def main() -> None:
    OUT_DIR.mkdir(parents=True, exist_ok=True)

    # rows[i] holds tuples for section i
    rows_per_section: list[list[tuple]] = [[] for _ in SECTION_FILES]

    files = sorted(p for p in DATA_DIR.iterdir()
                   if p.suffix in (".xls", ".xlsx") and not p.name.startswith("."))

    for path in files:
        year = parse_year(path.name)
        pairs = [pick_pair(r) for r in read_sheet_rows(path)]
        sections = split_sections(pairs)
        if len(sections) != 4:
            print(f"WARNING: {path.name} has {len(sections)} sections, expected 4")
            continue
        for i, section in enumerate(sections):
            title, data = clean_section(section)
            for orden, (cat, val) in enumerate(data, start=1):
                rows_per_section[i].append(
                    (year, title, orden, cat, coerce_value(val))
                )

    for i, fname in enumerate(SECTION_FILES):
        out_path = OUT_DIR / fname
        with out_path.open("w", newline="", encoding="utf-8") as fh:
            w = csv.writer(fh)
            w.writerow(["año", "titulo_seccion", "orden", "categoria", "valor"])
            w.writerows(rows_per_section[i])
        print(f"wrote {out_path}  ({len(rows_per_section[i])} rows)")


if __name__ == "__main__":
    main()
