"""
Normaliza los tabulados de migración de los Censos de Población y Vivienda
(1990, 2000, 2010, 2020) a una sola tabla larga estatal.

Cada censo publica el mismo par de tabulados con idéntico layout lógico:
  · "Migración 2" — población total por entidad de residencia actual y LUGAR DE
    NACIMIENTO  → stock (concepto='nacimiento')
  · "Migración 4" — población de 5 años y más por entidad de residencia actual y
    LUGAR DE RESIDENCIA 5 AÑOS ANTES → flujo (concepto='residencia_5a')

Sólo estos 4 censos preguntan por la residencia de 5 años antes. 1950 y 1960 no
tienen la pregunta; 1970 y 1980 usan "tiempo de residencia", que no es equivalente.

Fuentes: data/inegi/ccpv/migracion/
Salida:  dashboard_data/ccpv_migracion_estatal.parquet
         dashboard_data/ccpv_extranjeros_pais_2020.parquet

Run: uv run python scripts/prepare_ccpv_migracion.py
"""

import sys
import warnings
from pathlib import Path

import openpyxl
import pandas as pd
import polars as pl

sys.path.insert(0, str(Path(__file__).resolve().parent / "centralismo"))
from comun import NOMBRE, normalizar_estado

RAIZ = Path(__file__).resolve().parents[1]
DIR = RAIZ / "data" / "inegi" / "ccpv" / "migracion"
OUT_DIR = RAIZ / "dashboard_data"

# (censo, archivo, hoja, concepto, año_ref, offset de columna)
# offset=1 en 1990/2000: la columna 0 es la constante 'Estados Unidos Mexicanos'.
SPECS = [
    (1990, "CPyV90_Nal_Migracion.xlsx",         "CPyV90_Nal_MIG1",   "nacimiento",    None, 1),
    (1990, "CPyV90_Nal_Migracion.xlsx",         "CPyV90_Nal_MIG2",   "residencia_5a", 1985, 1),
    (2000, "CPyV2000_NAL_Migracion.xlsx",       "CPyV2000_Nal_MIG1", "nacimiento",    None, 1),
    (2000, "CPyV2000_NAL_Migracion.xlsx",       "CPyV2000_Nal_MIG3", "residencia_5a", 1995, 1),
    (2005, "Cont2005_NAL_Migracion.xls",        "Cont2005_Nal_MIG1", "residencia_5a", 2000, 1),
    (2010, "04_02B_ESTATAL.xls",                None,                "nacimiento",    None, 0),
    (2010, "04_04B_ESTATAL.xls",                None,                "residencia_5a", 2005, 0),
    (2020, "cpv2020_b_eum_04_migracion.xlsx",   "02",                "nacimiento",    None, 0),
    (2020, "cpv2020_b_eum_04_migracion.xlsx",   "04",                "residencia_5a", 2015, 0),
]

# Las etiquetas de categoría cambian de redacción en cada censo.
CATEGORIA = {
    "Total":                            "Total",
    "Población total":                  "Total",
    "Total residentes":                 "Total",
    "En la entidad":                    "En la entidad",
    "Nacidos en la entidad":            "En la entidad",
    "Residentes en la entidad":         "En la entidad",
    "En otra entidad":                  "En otra entidad",
    "Nacidos en otra entidad":          "En otra entidad",
    "Residentes en otra entidad":       "En otra entidad",
    "En los Estados Unidos de América": "En los Estados Unidos de América",
    "En Estados Unidos de América":     "En los Estados Unidos de América",
    "En otro país":                     "En otro país",
    "Nacidos en otro país":             "En otro país",
    "Residentes en otro país":          "En otro país",
    "No especificado":                  "No especificado",
}

NACIONAL = {"Estados Unidos Mexicanos", "Total Nacional"}

# Categoría exclusiva de leer_intercensal2015_agregado(): la Encuesta Intercensal
# 2015 no trae matriz origen-destino, así que el total de emigrantes no se puede
# derivar transponiendo el desglose (como en las demás fuentes) — se guarda como
# su propia categoría, fuera de la partición Total = En la entidad + En otra
# entidad + ... (no se debe sumar junto con ésas).
EMIGRANTES_AGREGADO = "Emigrantes (agregado)"


def _entero(v):
    """Celda → int, o None si no es un conteo (encabezado, nota al pie, centinela)."""
    if isinstance(v, bool) or v is None:
        return None
    if isinstance(v, int):
        return v
    if isinstance(v, float):
        return int(v) if v.is_integer() else None
    s = str(v).strip().replace(" ", "").replace(",", "").replace("\xa0", "")
    return int(s) if s.isdigit() else None


def _cve(etiqueta):
    """'09 Distrito Federal' → 9. Nacional → 0. 'Total'/vacío → None.

    El prefijo numérico 01–32 es estable en las 4 ediciones; los nombres no
    ('Distrito Federal'→'Ciudad de México', 'Veracruz - Llave', etc.).
    """
    if etiqueta is None:
        return None
    s = str(etiqueta).strip()
    if s in NACIONAL:
        return 0
    if len(s) > 2 and s[:2].isdigit():
        cve = int(s[:2])
        if 1 <= cve <= 32:
            return cve
    return normalizar_estado(s)


def _filas(archivo, hoja):
    """Devuelve las filas crudas del tabulado como listas de celdas."""
    ruta = DIR / archivo
    if ruta.suffix == ".xls":  # BIFF8: openpyxl no lo abre
        kwargs = {} if hoja is None else {"sheet_name": hoja}
        df = pd.read_excel(ruta, header=None, dtype=object, **kwargs)
        return df.values.tolist()
    with warnings.catch_warnings():
        warnings.simplefilter("ignore")
        wb = openpyxl.load_workbook(ruta, read_only=True, data_only=True)
        try:
            return [list(r) for r in wb[hoja].iter_rows(values_only=True)]
        finally:
            wb.close()


def leer_tabulado(censo, archivo, hoja, concepto, año_ref, off):
    filas = []
    desconocidas = set()
    for fila in _filas(archivo, hoja):
        celdas = list(fila[off:off + 6]) + [None] * 6
        destino, categoria, origen, personas, hombres, mujeres = celdas[:6]

        personas = _entero(personas)
        if personas is None:  # encabezados, banners, centinela '!', notas al pie
            continue
        cve_destino = _cve(destino)
        if cve_destino is None:
            continue

        clave = str(categoria).strip() if categoria is not None else ""
        if clave not in CATEGORIA:
            desconocidas.add(clave)
            continue

        # origen == 'Total' marca la fila agregada de la categoría; el resto son el
        # desglose por entidad de origen. Sólo 1990 añade 'Entidad federativa ins.
        # esp.' (entidad de origen insuficientemente especificada).
        etiq_origen = str(origen).strip() if origen is not None else "Total"
        cve_origen = _cve(etiq_origen)
        filas.append({
            "censo": censo,
            "concepto": concepto,
            "año_ref": año_ref,
            "cve_destino": cve_destino,
            "destino": NOMBRE.get(cve_destino, "Nacional"),
            "categoria": CATEGORIA[clave],
            "total_categoria": etiq_origen == "Total",
            "cve_origen": cve_origen,
            "origen": NOMBRE[cve_origen] if cve_origen else (
                None if etiq_origen == "Total" else "Entidad no especificada"
            ),
            "personas": personas,
            "hombres": _entero(hombres),
            "mujeres": _entero(mujeres),
        })

    if desconocidas:
        raise ValueError(
            f"{archivo}/{hoja}: categorías no reconocidas {sorted(desconocidas)}"
        )

    df = pl.DataFrame(filas, schema={
        "censo": pl.Int16, "concepto": pl.Utf8, "año_ref": pl.Int16,
        "cve_destino": pl.Int16, "destino": pl.Utf8, "categoria": pl.Utf8,
        "total_categoria": pl.Boolean, "cve_origen": pl.Int16, "origen": pl.Utf8,
        "personas": pl.Int64, "hombres": pl.Int64, "mujeres": pl.Int64,
    })
    print(f"  {censo} {concepto:<14} {archivo:<32} {df.height:>6,} filas")
    return df


def leer_paises_2020():
    """Origen-destino municipal 2020 → flujo desde el extranjero por país y estado.

    Trampas: el código de entidad de origen es de 3 dígitos ('001 Aguascalientes')
    y el de destino de 2 ('01 Aguascalientes'); código y nombre van pegados en un
    solo string. Los códigos >= 100 son países/continentes (997/998/999 = no
    especificado) y llevan '000 Total' en la columna de municipio.
    """
    ruta = DIR / "cpv2020_c_eum_origen_destino.xlsx"
    with warnings.catch_warnings():
        warnings.simplefilter("ignore")
        wb = openpyxl.load_workbook(ruta, read_only=True, data_only=True)
        filas = []
        for origen, _, destino, _, personas in wb["Origen-Destino"].iter_rows(
            min_row=8, max_col=5, values_only=True
        ):
            personas = _entero(personas)
            if personas is None or not isinstance(origen, str):
                continue
            cod, _, pais = origen.partition(" ")
            if not cod.isdigit() or int(cod) < 100:  # 001-032 = entidades
                continue
            cve_destino = _cve(destino)
            if cve_destino is None:
                continue
            filas.append({
                "cve_destino": cve_destino,
                "destino": NOMBRE[cve_destino],
                "cod_pais": int(cod),
                "pais": pais.strip(),
                "personas": personas,
            })
        wb.close()

    df = (
        pl.DataFrame(filas, schema={
            "cve_destino": pl.Int16, "destino": pl.Utf8,
            "cod_pais": pl.Int16, "pais": pl.Utf8, "personas": pl.Int64,
        })
        .group_by(["cve_destino", "destino", "cod_pais", "pais"])
        .agg(pl.col("personas").sum())
        .sort(["cve_destino", "personas"], descending=[False, True])
    )
    print(f"  2020 países      cpv2020_c_eum_origen_destino.xlsx {df.height:>6,} filas")
    return df


def leer_intercensal2015_agregado():
    """04_migracion.xls, hoja "06" — Encuesta Intercensal 2015.

    A diferencia de los censos, esta encuesta MUESTRAL no publica la matriz
    origen-destino por entidad: sólo trae, por entidad, población/inmigrantes/
    emigrantes/saldo agregados (fila Estimador == 'Valor'). Se homologa al
    esquema de leer_tabulado() con 3 filas total_categoria por entidad (Total /
    En otra entidad / EMIGRANTES_AGREGADO) y CERO filas de desglose por origen.
    """
    ruta = DIR / "04_migracion.xls"
    filas = []
    datos = pd.read_excel(ruta, sheet_name="06", header=None, dtype=object).values.tolist()
    for fila in datos:
        entidad, sexo, estimador = fila[0], fila[1], fila[2]
        if sexo != "Total" or estimador != "Valor":
            continue
        cve = _cve(entidad)
        if cve is None:
            continue
        poblacion, inm, emi = _entero(fila[3]), _entero(fila[4]), _entero(fila[5])
        for categoria, valor in [("Total", poblacion), ("En otra entidad", inm),
                                  (EMIGRANTES_AGREGADO, emi)]:
            filas.append({
                "censo": 2015, "concepto": "residencia_5a", "año_ref": 2010,
                "cve_destino": cve, "destino": NOMBRE.get(cve, "Nacional"),
                "categoria": categoria, "total_categoria": True,
                "cve_origen": None, "origen": None,
                "personas": valor, "hombres": None, "mujeres": None,
            })

    df = pl.DataFrame(filas, schema={
        "censo": pl.Int16, "concepto": pl.Utf8, "año_ref": pl.Int16,
        "cve_destino": pl.Int16, "destino": pl.Utf8, "categoria": pl.Utf8,
        "total_categoria": pl.Boolean, "cve_origen": pl.Int16, "origen": pl.Utf8,
        "personas": pl.Int64, "hombres": pl.Int64, "mujeres": pl.Int64,
    })
    print(f"  2015 residencia_5a    04_migracion.xls (hoja 06, agregado) {df.height:>6,} filas")
    return df


def validar(mig, paises):
    for (censo, concepto), g in mig.group_by(["censo", "concepto"], maintain_order=True):
        estados = g.filter(pl.col("cve_destino") > 0)
        assert estados["cve_destino"].n_unique() == 32, \
            f"{censo}/{concepto}: {estados['cve_destino'].n_unique()} entidades destino"

        agregado = g.filter(pl.col("total_categoria"))
        desglose = g.filter(~pl.col("total_categoria"))

        if desglose.height == 0:
            # Fuente agregada sin matriz origen-destino (Encuesta Intercensal
            # 2015): sólo se valida con las cifras nacionales de `esperado` abajo.
            continue

        # a) Las categorías deben sumar la fila 'Total' de cada entidad.
        partes = (
            agregado.filter(pl.col("categoria") != "Total")
            .group_by("cve_destino").agg(pl.col("personas").sum().alias("suma"))
        )
        total = agregado.filter(pl.col("categoria") == "Total").select(
            ["cve_destino", pl.col("personas").alias("total")]
        )
        d = partes.join(total, on="cve_destino").filter(pl.col("suma") != pl.col("total"))
        assert d.height == 0, f"{censo}/{concepto}: categorías no suman\n{d}"

        # b) El desglose por origen debe sumar el agregado de 'En otra entidad'.
        partes = desglose.group_by("cve_destino").agg(pl.col("personas").sum().alias("suma"))
        total = agregado.filter(pl.col("categoria") == "En otra entidad").select(
            ["cve_destino", pl.col("personas").alias("total")]
        )
        d = partes.join(total, on="cve_destino").filter(pl.col("suma") != pl.col("total"))
        assert d.height == 0, f"{censo}/{concepto}: desglose no suma\n{d}"

        # c) Transponer la matriz debe reproducir el bloque nacional por origen.
        emig = (
            desglose.filter((pl.col("cve_destino") > 0) & pl.col("cve_origen").is_not_null())
            .group_by("cve_origen").agg(pl.col("personas").sum().alias("suma"))
        )
        nal = desglose.filter(
            (pl.col("cve_destino") == 0) & pl.col("cve_origen").is_not_null()
        ).select(["cve_origen", pl.col("personas").alias("total")])
        d = emig.join(nal, on="cve_origen").filter(pl.col("suma") != pl.col("total"))
        assert d.height == 0 and emig.height == 32, \
            f"{censo}/{concepto}: matriz interestatal descuadrada\n{d}"

    # 2020: el flujo extranjero del tabulado 04 y el del origen-destino deben coincidir.
    tab = mig.filter(
        (pl.col("censo") == 2020) & (pl.col("concepto") == "residencia_5a")
        & (pl.col("cve_destino") > 0) & pl.col("total_categoria")
        & pl.col("categoria").is_in(["En los Estados Unidos de América", "En otro país"])
    )["personas"].sum()
    od = paises.filter(pl.col("cod_pais") < 997)["personas"].sum()
    assert abs(tab - od) / tab < 0.01, f"flujo extranjero 2020: tabulado {tab:,} vs O-D {od:,}"
    print(f"  flujo extranjero 2020: tabulado {tab:,} · origen-destino {od:,}")

    # Cifras nacionales publicadas por INEGI.
    def nacional(censo, concepto, categoria):
        return mig.filter(
            (pl.col("censo") == censo) & (pl.col("concepto") == concepto)
            & (pl.col("cve_destino") == 0) & (pl.col("categoria") == categoria)
            & pl.col("total_categoria")
        )["personas"].sum()

    esperado = {
        (2020, "residencia_5a", "Total"):           115_693_273,
        (2020, "residencia_5a", "En otra entidad"):   3_807_844,
        (2020, "nacimiento",    "En otra entidad"):  21_611_963,
        (2020, "nacimiento", "En los Estados Unidos de América"): 797_266,
        (2020, "nacimiento",    "En otro país"):        414_986,
        (2010, "residencia_5a", "En otra entidad"):   3_292_310,
        (2000, "residencia_5a", "En otra entidad"):   3_584_957,
        (1990, "residencia_5a", "En otra entidad"):   3_477_237,
        (2005, "residencia_5a", "Total"):             90_266_425,
        (2005, "residencia_5a", "En otra entidad"):    2_410_407,
        (2015, "residencia_5a", "Total"):             107_396_355,
        (2015, "residencia_5a", "En otra entidad"):     3_197_619,
    }
    for (censo, concepto, cat), val in esperado.items():
        obt = nacional(censo, concepto, cat)
        assert obt == val, f"{censo}/{concepto}/{cat}: {obt:,} != {val:,}"
    print(f"  {len(esperado)} cifras nacionales de INEGI verificadas")

    # 2015: sin matriz, inmigrantes nacionales debe igualar emigrantes nacionales
    # (por construcción — es la misma migración interna vista desde ambos lados).
    inm_nal = nacional(2015, "residencia_5a", "En otra entidad")
    emi_nal = nacional(2015, "residencia_5a", EMIGRANTES_AGREGADO)
    assert inm_nal == emi_nal == 3_197_619, f"2015: inmigrantes {inm_nal:,} != emigrantes {emi_nal:,}"
    print(f"  2015: inmigrantes = emigrantes = {inm_nal:,} (auto-consistencia nacional)")


def main():
    OUT_DIR.mkdir(parents=True, exist_ok=True)
    print("Tabulados estatales:")
    mig = pl.concat([leer_tabulado(*s) for s in SPECS] + [leer_intercensal2015_agregado()])
    print("Desglose por país:")
    paises = leer_paises_2020()

    print("Validación:")
    validar(mig, paises)

    for nombre, df in [("ccpv_migracion_estatal.parquet", mig),
                       ("ccpv_extranjeros_pais_2020.parquet", paises)]:
        df.write_parquet(OUT_DIR / nombre)
        print(f"Saved → {OUT_DIR / nombre}  ({df.height:,} filas)")


if __name__ == "__main__":
    main()
