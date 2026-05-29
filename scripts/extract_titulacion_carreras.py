"""Extract títulos expedidos por carrera/opción from every UNAM yearbook in
data/unam_titulacion_carreras/ and concatenate them into a single CSV.

Each yearbook file is named YYYY_titulo.xls(x) and contains one sheet with a
three-level hierarchy:

    Entidad académica            (Facultad / Escuela / Centro / Instituto / ...)
      Carrera                    (e.g. Arquitectura)
        Opción de titulación     (e.g. Tesis o tesina y examen profesional)

The script emits one row per (entidad, carrera, opción) with hombres / mujeres /
total counts. The year is taken from the filename.
"""

import csv
import re
from pathlib import Path

import openpyxl
import xlrd

DATA_DIR = Path("data/unam/titulacion_carreras")
OUT_PATH = Path("data/unam_egresos_y_titulacion/titulacion_carreras.csv")

ENTIDAD_PREFIXES = ("Facultad", "Escuela", "Centro", "Instituto", "Programa", "Unidad")

# Known opciones de titulación across all years (the leaf level of the hierarchy).
# Variants (encoding, case, wording) are mapped to a canonical name.
METHOD_CANON = {
    "Tesis o tesina y examen profesional": "Tesis o tesina y examen profesional",
    "Seminario de tesis o tesina": "Seminario de tesis o tesina",
    "Ampliación y profundización de conocimientos": "Ampliación y profundización de conocimientos",
    "Trabajo profesional": "Trabajo profesional",
    "Estudios de posgrado": "Estudios de posgrado",
    "Estudios en posgrado": "Estudios de posgrado",
    "Créditos y alto nivel académico": "Créditos y alto nivel académico",
    "Actividad de investigación": "Actividad de investigación",
    "Actividad de investigaciÃ³n": "Actividad de investigación",
    "Actividad de apoyo a la docencia": "Actividad de apoyo a la docencia",
    "Servicio social": "Servicio social",
    "Examen general de conocimientos": "Examen general de conocimientos",
    "Examen General de conocimientos": "Examen general de conocimientos",
    "Otra": "Otra",
    "Otras": "Otra",
}

FOOTER_PREFIX = "Para mayor información"


def read_sheet_rows(path: Path) -> list[tuple]:
    """Return rows of the titulación sheet as (col0, col1, col2, col3) tuples."""
    if path.suffix == ".xls":
        wb = xlrd.open_workbook(str(path))
        ws = wb.sheet_by_index(0)
        return [
            tuple(ws.cell_value(r, c) if c < ws.ncols else None for c in range(4))
            for r in range(ws.nrows)
        ]
    wb = openpyxl.load_workbook(str(path), data_only=True, read_only=True)
    ws = wb[wb.sheetnames[0]]
    out: list[tuple] = []
    for row in ws.iter_rows(values_only=True):
        r = list(row) + [None] * 4
        out.append((r[0], r[1], r[2], r[3]))
    return out


def coerce_int(v) -> int | None:
    if v is None:
        return None
    if isinstance(v, str):
        s = v.strip()
        if not s or s == "-":
            return None
        try:
            return int(float(s))
        except ValueError:
            return None
    if isinstance(v, (int, float)):
        return int(v)
    return None


def parse_year(filename: str) -> int:
    m = re.match(r"(\d{4})", filename)
    if not m:
        raise ValueError(f"cannot parse year from {filename}")
    return int(m.group(1))


def normalize_entidad(s: str) -> str:
    # drop "(continuación)" suffix used when a section spans pages
    return re.sub(r"\s*\(continuación\)\s*$", "", s).strip()


def extract_rows(path: Path) -> list[tuple]:
    """Yield (año, entidad, carrera, opcion_titulacion, hombres, mujeres, total)."""
    year = parse_year(path.name)
    raw = read_sheet_rows(path)

    entidad = ""
    carrera = ""
    header_seen = False
    out: list[tuple] = []

    for a, b, c, d in raw:
        s = "" if a is None else str(a).strip()
        if not s:
            continue
        if s.startswith(FOOTER_PREFIX):
            continue

        # column header row: ('Entidad académica / Carrera / Opción de titulación', 'Hombres', ...)
        if not header_seen:
            if s.lower().startswith("entidad"):
                header_seen = True
            continue

        # skip aggregate rows that aren't part of the hierarchy
        if s.upper().startswith("T O T A L"):
            continue
        if s == "Licenciatura":  # appears once in 2019 as a duplicate grand total
            continue

        if s.startswith(ENTIDAD_PREFIXES):
            entidad = normalize_entidad(s)
            carrera = ""
            continue

        canon = METHOD_CANON.get(s)
        if canon is not None:
            out.append(
                (year, entidad, carrera, canon,
                 coerce_int(b), coerce_int(c), coerce_int(d))
            )
        else:
            carrera = s

    return out


def main() -> None:
    OUT_PATH.parent.mkdir(parents=True, exist_ok=True)
    all_rows: list[tuple] = []
    files = sorted(p for p in DATA_DIR.iterdir()
                   if p.suffix in (".xls", ".xlsx") and not p.name.startswith("."))
    for path in files:
        all_rows.extend(extract_rows(path))

    with OUT_PATH.open("w", newline="", encoding="utf-8") as fh:
        w = csv.writer(fh)
        w.writerow(["año", "entidad", "carrera", "opcion_titulacion",
                    "hombres", "mujeres", "total"])
        w.writerows(all_rows)
    print(f"wrote {OUT_PATH}  ({len(all_rows)} rows from {len(files)} files)")


if __name__ == "__main__":
    main()
