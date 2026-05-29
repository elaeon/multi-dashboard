"""Extract egreso by carrera from every UNAM yearbook in data/unam/egreso_carreras/.

Each file is named egreso_YYYY.xls(x) and contains one sheet with a two-level
hierarchy: Entidad académica → Carrera, with columns Hombres / Mujeres / Total.
"""

import csv
import re
from pathlib import Path

import openpyxl
import xlrd

DATA_DIR = Path("data/unam/egreso_carreras")
OUT_PATH = Path("data/unam_egresos_y_titulacion/egreso_carreras.csv")

ENTIDAD_PREFIXES = (
    "Facultad", "Escuela", "Centro", "Instituto", "Programa", "Unidad", "Colegio",
)
SKIP_PREFIXES = ("T O T A L", "Para mayor", "FUENTE", "a Las", "b Las", "UNAM.")


def read_rows(path: Path) -> list[list]:
    if path.suffix == ".xls":
        wb = xlrd.open_workbook(str(path))
        ws = wb.sheet_by_index(0)
        return [
            [ws.cell_value(r, c) if c < ws.ncols else None for c in range(4)]
            for r in range(ws.nrows)
        ]
    wb = openpyxl.load_workbook(str(path), data_only=True, read_only=True)
    ws = wb[wb.sheetnames[0]]
    return [
        [(list(row) + [None] * 4)[c] for c in range(4)]
        for row in ws.iter_rows(values_only=True)
    ]


def find_header_row(rows: list[list]) -> int:
    """Return index of the row containing 'Entidad académica'."""
    for i, row in enumerate(rows):
        if row[0] and "Entidad" in str(row[0]):
            return i
    raise ValueError("Column header row not found")


def coerce_int(v) -> int:
    if v in (None, "-", "", " "):
        return 0
    try:
        return int(float(v))
    except (ValueError, TypeError):
        return 0


def parse_year(filename: str) -> int:
    m = re.search(r"(\d{4})", filename)
    if not m:
        raise ValueError(f"Cannot parse year from {filename}")
    return int(m.group(1))


def extract(path: Path) -> list[tuple]:
    year = parse_year(path.name)
    rows = read_rows(path)
    hdr = find_header_row(rows)

    records = []
    entidad = ""
    for row in rows[hdr + 1:]:
        name = str(row[0]).strip() if row[0] is not None else ""
        if not name:
            continue
        if any(name.startswith(p) for p in SKIP_PREFIXES):
            continue
        if name.startswith(ENTIDAD_PREFIXES):
            entidad = name
            continue
        records.append((
            year, entidad, name,
            coerce_int(row[1]), coerce_int(row[2]), coerce_int(row[3]),
        ))
    return records


def main() -> None:
    OUT_PATH.parent.mkdir(parents=True, exist_ok=True)
    all_rows: list[tuple] = []
    files = sorted(
        p for p in DATA_DIR.iterdir()
        if p.name.startswith("egreso") and p.suffix in (".xls", ".xlsx")
    )
    for path in files:
        rows = extract(path)
        all_rows.extend(rows)
        print(f"  {path.name}: {len(rows)} rows")

    with OUT_PATH.open("w", newline="", encoding="utf-8") as fh:
        w = csv.writer(fh)
        w.writerow(["año", "entidad", "carrera", "hombres", "mujeres", "total"])
        w.writerows(all_rows)
    print(f"\nWrote {OUT_PATH}  ({len(all_rows)} rows from {len(files)} files)")


if __name__ == "__main__":
    main()
