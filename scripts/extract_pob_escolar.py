"""Extract the 'pob_escolar' sheet from every UNAM yearbook in data/unam/
and concatenate it into a single CSV.

The sheet has the structure:
    UNAM
    POBLACIÓN ESCOLAR TOTAL
    <ciclo>                       (e.g. 2024-2025)
    (blank)  | Primer Ingreso | Reingreso | Total
    Posgrado | ...
      Sistema Escolarizado ...
      Sistema Universidad Abierta ...
    Licenciatura ...
      ...
    Técnico Profesional ...
    Bachillerato ...
      ...
    T O T A L
    Para mayor información ...

The year is taken from the filename (YYYY.xls(x)).
"""

import csv
import re
from pathlib import Path

import openpyxl
import xlrd

DATA_DIR = Path("data/unam")
OUT_PATH = Path("data/unam_poblacion_escolar.csv")

PARENT_LEVELS = {"Posgrado", "Licenciatura", "Técnico Profesional", "Bachillerato"}
FOOTER_PREFIX = "Para mayor información"


def read_sheet_rows(path: Path) -> list[list]:
    """Return rows of the pob_escolar sheet as [col0, col1, col2, col3]."""
    if path.suffix == ".xls":
        wb = xlrd.open_workbook(str(path))
        name = next(n for n in wb.sheet_names() if "pob" in n.lower())
        ws = wb.sheet_by_name(name)
        return [[ws.cell_value(r, c) if c < ws.ncols else None for c in range(4)]
                for r in range(ws.nrows)]

    wb = openpyxl.load_workbook(str(path), data_only=True, read_only=True)
    name = next(n for n in wb.sheetnames if "pob" in n.lower())
    ws = wb[name]
    return [[row[c] if c < len(row) else None for c in range(4)]
            for row in ws.iter_rows(values_only=True)]


def coerce_int(v) -> int | None:
    """Cast numeric cell to int; '-' and blank → None."""
    if v is None:
        return None
    if isinstance(v, str):
        s = v.strip()
        if not s or s == "-":
            return None
        try:
            return int(float(s))
        except ValueError:
            return None
    if isinstance(v, (int, float)):
        return int(v)
    return None


def parse_year(filename: str) -> int:
    m = re.match(r"(\d{4})", filename)
    if not m:
        raise ValueError(f"cannot parse year from {filename}")
    return int(m.group(1))


def extract_rows(path: Path) -> list[tuple]:
    """Yield (año, ciclo, nivel_padre, categoria, primer_ingreso, reingreso, total)."""
    year = parse_year(path.name)
    raw = read_sheet_rows(path)

    ciclo = ""
    header_seen = False
    parent: str | None = None
    out: list[tuple] = []

    for a, b, c, d in raw:
        a_str = "" if a is None else str(a).strip()
        b_str = "" if b is None else str(b).strip()

        # capture ciclo (e.g. "2024-2025") — appears before the column header row
        if not header_seen and re.fullmatch(r"\d{4}-\d{4}", a_str):
            ciclo = a_str
            continue

        # column header row: ('', 'Primer Ingreso', 'Reingreso', 'Total')
        if not header_seen and b_str.lower().startswith("primer"):
            header_seen = True
            continue

        if not header_seen:
            continue
        if not a_str:
            continue
        if a_str.startswith(FOOTER_PREFIX):
            continue

        pi, re_, tot = coerce_int(b), coerce_int(c), coerce_int(d)

        if a_str.upper() == "T O T A L":
            out.append((year, ciclo, "", a_str, pi, re_, tot))
            continue

        if a_str in PARENT_LEVELS:
            parent = a_str
            out.append((year, ciclo, "", a_str, pi, re_, tot))
        else:
            out.append((year, ciclo, parent or "", a_str, pi, re_, tot))

    return out


def main() -> None:
    OUT_PATH.parent.mkdir(parents=True, exist_ok=True)
    all_rows: list[tuple] = []
    files = sorted(p for p in DATA_DIR.iterdir()
                   if p.suffix in (".xls", ".xlsx") and not p.name.startswith("."))
    for path in files:
        all_rows.extend(extract_rows(path))

    with OUT_PATH.open("w", newline="", encoding="utf-8") as fh:
        w = csv.writer(fh)
        w.writerow(["año", "ciclo", "nivel_padre", "categoria",
                    "primer_ingreso", "reingreso", "total"])
        w.writerows(all_rows)
    print(f"wrote {OUT_PATH}  ({len(all_rows)} rows from {len(files)} files)")


if __name__ == "__main__":
    main()
