"""Extract primer ingreso and reingreso by carrera from all licenciatura
population files in data/unam/poblacion_escolar/.

Column layout varies slightly across years, so column positions are detected
dynamically from the sub-header row that contains "Hombres"/"Mujeres" labels.
"""

import csv
import re
from pathlib import Path

import openpyxl
import xlrd

DATA_DIR = Path("data/unam/poblacion_escolar")
OUT_PATH = Path("data/unam_egresos_y_titulacion/poblacion_escolar_lic.csv")

ENTIDAD_PREFIXES = (
    "Facultad", "Escuela", "Centro", "Instituto", "Programa", "Unidad", "Colegio",
)
SKIP_PREFIXES = ("T O T A L", "Para mayor", "FUENTE", "a Las", "UNAM.", "b ")


def read_rows(path: Path) -> list[list]:
    """Return all rows as lists of up-to-9 values."""
    if path.suffix == ".xls":
        wb = xlrd.open_workbook(str(path))
        ws = wb.sheet_by_index(0)
        return [
            [ws.cell_value(r, c) if c < ws.ncols else None for c in range(9)]
            for r in range(ws.nrows)
        ]
    wb = openpyxl.load_workbook(str(path), data_only=True, read_only=True)
    ws = wb[wb.sheetnames[0]]
    return [
        [(list(row) + [None] * 9)[c] for c in range(9)]
        for row in ws.iter_rows(values_only=True)
    ]


def detect_columns(rows: list[list]) -> tuple[int, int, int, int, int]:
    """Return (header_row_idx, pi_h_col, pi_m_col, rei_h_col, rei_m_col).

    Scans for the sub-header row where 'Hombres' appears twice.
    First occurrence → primer ingreso; second → reingreso.
    """
    for i, row in enumerate(rows):
        stripped = [str(v).strip() if v is not None else "" for v in row]
        hombres_cols = [j for j, v in enumerate(stripped) if "Hombres" in v or "hombres" in v]
        mujeres_cols = [j for j, v in enumerate(stripped) if "Mujeres" in v or "mujeres" in v]
        if len(hombres_cols) >= 2 and len(mujeres_cols) >= 2:
            return i, hombres_cols[0], mujeres_cols[0], hombres_cols[1], mujeres_cols[1]
    raise ValueError("Sub-header row with two 'Hombres' columns not found")


def coerce_int(v) -> int:
    if v in (None, "-", "", " "):
        return 0
    try:
        return int(float(v))
    except (ValueError, TypeError):
        return 0


def parse_year(filename: str) -> int:
    m = re.search(r"(\d{4})", filename)
    if not m:
        raise ValueError(f"Cannot parse year from {filename}")
    return int(m.group(1))


def extract(path: Path) -> list[tuple]:
    year = parse_year(path.name)
    rows = read_rows(path)
    hdr_idx, pi_h_col, pi_m_col, rei_h_col, rei_m_col = detect_columns(rows)

    records = []
    entidad = ""
    for row in rows[hdr_idx + 1 :]:
        name = str(row[0]).strip() if row[0] is not None else ""
        if not name:
            continue
        if any(name.startswith(p) for p in SKIP_PREFIXES):
            continue
        if name.startswith(ENTIDAD_PREFIXES):
            entidad = name
            continue
        pi_h = coerce_int(row[pi_h_col])
        pi_m = coerce_int(row[pi_m_col])
        rei_h = coerce_int(row[rei_h_col])
        rei_m = coerce_int(row[rei_m_col])
        records.append((year, entidad, name, pi_h, pi_m, rei_h, rei_m))
    return records


def main() -> None:
    OUT_PATH.parent.mkdir(parents=True, exist_ok=True)
    all_rows: list[tuple] = []
    files = sorted(
        p for p in DATA_DIR.iterdir()
        if p.name.startswith("licenciatura") and p.suffix in (".xls", ".xlsx")
    )
    for path in files:
        rows = extract(path)
        all_rows.extend(rows)
        print(f"  {path.name}: {len(rows)} rows")

    with OUT_PATH.open("w", newline="", encoding="utf-8") as fh:
        w = csv.writer(fh)
        w.writerow(["año", "entidad", "carrera", "pi_h", "pi_m", "rei_h", "rei_m"])
        w.writerows(all_rows)
    print(f"\nWrote {OUT_PATH}  ({len(all_rows)} rows from {len(files)} files)")


if __name__ == "__main__":
    main()
